import gspread
from oauth2client.service_account import ServiceAccountCredentials
import pandas as pd
import schedule
import time
from openpyxl import load_workbook
from openpyxl import Workbook

scope = ['https://www.googleapis.com/auth/spreadsheets']
credentials = ServiceAccountCredentials.from_json_keyfile_name('spropusk-1c0b9f20a2f2.json', scope)

client = gspread.authorize(credentials)

spreadsheet = client.open_by_url('https://docs.google.com/spreadsheets/d/1GROzJcUJNWG2KST-q3opt4iq39Cl12VeGj2uUpT6lCQ/edit?pli=1#gid=0')

worksheet = spreadsheet.worksheet('БИГ ДЭЙТА')
column_names = worksheet.row_values(1)
def load_data_from_google_sheet1(credentials_file, excel_file):

    workbook = load_workbook('Тогумбвев Владислав.xlsx')
    worksheet = workbook['Sheet1']
    range_to_clear = 'A1:O999'  # Замените на нужный диапазон
    for row in worksheet[range_to_clear]:
        for cell in row:
            cell.value = None
    workbook.save('Тогумбвев Владислав.xlsx')

    scope = ["https://spreadsheets.google.com/feeds", "https://www.googleapis.com/auth/drive"]
    credentials = ServiceAccountCredentials.from_json_keyfile_name(credentials_file, scope)
    client = gspread.authorize(credentials)
    worksheet = client.open("ПРОПУСКА").worksheet('Тогумбаев Владислав')
    data = worksheet.get_all_values()
    df = pd.DataFrame(data[1:], columns=data[0])
    df.to_excel(excel_file, index=False)
def load_data_from_google_sheet2(credentials_file, excel_file):

    workbook = load_workbook(excel_file)
    sheet = workbook.active
    sheet.delete_rows(2, sheet.max_row)

    scope = ["https://spreadsheets.google.com/feeds", "https://www.googleapis.com/auth/drive"]
    credentials = ServiceAccountCredentials.from_json_keyfile_name(credentials_file, scope)
    client = gspread.authorize(credentials)
    worksheet = client.open("ПРОПУСКА").worksheet('Коптев Константин')
    data = worksheet.get_all_values()
    df = pd.DataFrame(data[1:], columns=data[0])
    df.to_excel(excel_file, index=False)
def load_data_from_google_sheet3(credentials_file, excel_file):
    workbook = load_workbook(excel_file)
    sheet = workbook.active
    sheet.delete_rows(2, sheet.max_row)

    scope = ["https://spreadsheets.google.com/feeds", "https://www.googleapis.com/auth/drive"]
    credentials = ServiceAccountCredentials.from_json_keyfile_name(credentials_file, scope)
    client = gspread.authorize(credentials)
    worksheet = client.open("ПРОПУСКА").worksheet('Кондратьев Павел')
    data = worksheet.get_all_values()
    df = pd.DataFrame(data[1:], columns=data[0])
    df.to_excel(excel_file, index=False)
def load_data_from_google_sheet4(credentials_file, excel_file):
    workbook = load_workbook(excel_file)
    sheet = workbook.active
    sheet.delete_rows(2, sheet.max_row)

    scope = ["https://spreadsheets.google.com/feeds", "https://www.googleapis.com/auth/drive"]
    credentials = ServiceAccountCredentials.from_json_keyfile_name(credentials_file, scope)
    client = gspread.authorize(credentials)
    worksheet = client.open("ПРОПУСКА").worksheet('Максим Стрелков')
    data = worksheet.get_all_values()
    df = pd.DataFrame(data[1:], columns=data[0])
    df.to_excel(excel_file, index=False)
def load_data_from_google_sheet_Big_Date(credentials_file, excel_file, filter_column1, filter_column2):
    filter_value1 = 'Готов '
    filter_value2 = 'Не оплачен'

    scope = ["https://www.googleapis.com/auth/spreadsheets", "https://www.googleapis.com/auth/drive"]
    credentials = ServiceAccountCredentials.from_json_keyfile_name(credentials_file, scope)
    client = gspread.authorize(credentials)

    spreadsheet = client.open("ПРОПУСКА")
    worksheet = spreadsheet.worksheet('БИГ ДЭЙТА')

    filtered_rows = worksheet.get_all_records()
    filtered_rows = [row for row in filtered_rows if row[filter_column1] == filter_value1 and row[filter_column2] == filter_value2]

    workbook = load_workbook(excel_file)
    sheet = workbook.active
    sheet.delete_rows(1, sheet.max_row)

    filtered_columns = list(filtered_rows[0].keys())
    sheet.append(filtered_columns)
    for row in filtered_rows:
        data = [row[column] for column in filtered_columns]
        sheet.append(data)
    workbook.save(excel_file)

filter_column1 = column_names[8]
filter_column2 = column_names[11]

def job():
    credentials_file = "spropusk-1c0b9f20a2f2.json"
    excel_file1 = "Тогумбвев Владислав.xlsx"
    excel_file2 = "Коптев Константин.xlsx"
    excel_file3 = "Кондратьев Павел.xlsx"
    excel_file4 = "Максим Стрелков.xlsx"
    excel_file_Big_Date = "БИГ ДЭЙТА.xlsx"
    load_data_from_google_sheet1(credentials_file, excel_file1)
    load_data_from_google_sheet2(credentials_file, excel_file2)
    load_data_from_google_sheet3(credentials_file, excel_file3)
    load_data_from_google_sheet4(credentials_file, excel_file4)
    load_data_from_google_sheet_Big_Date(credentials_file, excel_file_Big_Date, filter_column1, filter_column2)

schedule.every().day.at("10:00").do(job)

while True:
    schedule.run_pending()
    time.sleep(1)



