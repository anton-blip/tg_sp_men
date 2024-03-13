import telebot
import openpyxl
import re
import threading
import schedule
import time

bot_token = "6739766780:AAFITmpW94N6gYG3qPqQLjVVpuStDGcL9nE"

allowed_users = ["6707703501", "1222339745", "1699179440","6720840984", "6655526374", "397255972"]
admin_users = ["6707703501", "1222339745", "1699179440", "397255972"]
bot = telebot.TeleBot(bot_token)

@bot.message_handler(commands=['start'])
def start(message):
    chat_id = str(message.chat.id)
    if chat_id in allowed_users:
        bot.send_message(chat_id, "Вы успешно аутентифицированы!\n Вы можите испоьзовать /dolg")
    else:
        bot.send_message(chat_id, "У вас нет доступа к этому боту.")

@bot.message_handler(commands=['dolg'])
def meneger(message):
    chat_id = str(message.chat.id)
    if chat_id == "6707703501":
        file_path = 'БИГ ДЭЙТА.xlsx'
        with open(file_path, 'rb') as f:
            bot.send_document(message.chat.id, f)
        workbook = openpyxl.load_workbook('БИГ ДЭЙТА.xlsx')
        sheet = workbook['Лист1']
        total = 0
        total += sum(float(re.sub(r'р\.|\s|,00', '', row[6])) for row in sheet.iter_rows(min_row=2, values_only=True) if row[6] is not None)
        bot.send_message(chat_id,f'Сумма дебиторской задолженности (Общая): \n {'{:,}'.format(total).replace(',', ' ')} р.')

    if chat_id == "1699179440":
        file_path = 'Тогумбвев Владислав.xlsx'
        with open(file_path, 'rb') as f:
            bot.send_document(message.chat.id, f)

        workbook = openpyxl.load_workbook('Тогумбвев Владислав.xlsx')
        sheet = workbook['Sheet1']
        total = 0
        total += sum(float(row[6]) for row in sheet.iter_rows(min_row=2, values_only=True) if row[6] is not None)
        bot.send_message(chat_id, f'Сумма дебиторской задолженности: {'{:,}'.format(total).replace(',', ' ')} р.')

    if chat_id == "6720840984":
        file_path = 'Кондратьев Павел.xlsx'
        with open(file_path, 'rb') as f:
            bot.send_document(message.chat.id, f)

        workbook = openpyxl.load_workbook('Кондратьев Павел.xlsx')
        sheet = workbook['Sheet1']
        total = 0
        total += sum(float(row[6]) for row in sheet.iter_rows(min_row=2, values_only=True) if row[6] is not None)
        bot.send_message(chat_id, f'Сумма дебиторской задолженности: {'{:,}'.format(total).replace(',', ' ')} р.')

    if chat_id == "6655526374":
        file_path = 'Максим Стрелков.xlsx'
        with open(file_path, 'rb') as f:
            bot.send_document(message.chat.id, f)

        workbook = openpyxl.load_workbook('Максим Стрелков.xlsx')
        sheet = workbook['Sheet1']
        total = 0
        total += sum(float(row[6]) for row in sheet.iter_rows(min_row=2, values_only=True) if row[6] is not None)
        bot.send_message(chat_id, f'Сумма дебиторской задолженности: {'{:,}'.format(total).replace(',', ' ')} р.')

    elif not(chat_id in allowed_users):
        bot.send_message(chat_id, "У вас нет доступа к этомой функции.")
@bot.message_handler(commands=['admin'])
def message_admin(message):
    chat_id = str(message.chat.id)
    if chat_id in admin_users:
        file_path = 'Тогумбвев Владислав.xlsx'
        with open(file_path, 'rb') as f:
            bot.send_document(message.chat.id, f)
        workbook = openpyxl.load_workbook('Тогумбвев Владислав.xlsx')
        sheet = workbook['Sheet1']
        total = 0
        total += sum(float(row[6]) for row in sheet.iter_rows(min_row=2, values_only=True) if row[6] is not None)
        bot.send_message(chat_id, f'Сумма дебиторской задолженности (Клиентов Тогумбвев Владислав):\n {'{:,}'.format(total).replace(',', ' ')} р.')

        file_path = 'Максим Стрелков.xlsx'
        with open(file_path, 'rb') as f:
            bot.send_document(message.chat.id, f)
        workbook = openpyxl.load_workbook('Максим Стрелков.xlsx')
        sheet = workbook['Sheet1']
        total = 0
        total += sum(float(row[6]) for row in sheet.iter_rows(min_row=2, values_only=True) if row[6] is not None)
        bot.send_message(chat_id, f'Сумма дебиторской задолженности (Клиентов Максим Стрелков):\n {'{:,}'.format(total).replace(',', ' ')} р.')

        file_path = 'Кондратьев Павел.xlsx'
        with open(file_path, 'rb') as f:
            bot.send_document(message.chat.id, f)
        workbook = openpyxl.load_workbook('Кондратьев Павел.xlsx')
        sheet = workbook['Sheet1']
        total = 0
        total += sum(float(row[6]) for row in sheet.iter_rows(min_row=2, values_only=True) if row[6] is not None)
        bot.send_message(chat_id, f'Сумма дебиторской задолженности (Клиентов Кондратьев Павел):\n {'{:,}'.format(total).replace(',', ' ')} р.')

        file_path = 'Коптев Константин.xlsx'
        with open(file_path, 'rb') as f:
            bot.send_document(message.chat.id, f)
        workbook = openpyxl.load_workbook('Коптев Константин.xlsx')
        sheet = workbook['Sheet1']
        total = 0
        total += sum(float(row[6]) for row in sheet.iter_rows(min_row=2, values_only=True) if row[6] is not None)
        bot.send_message(chat_id, f'Сумма дебиторской задолженности (Клиентов Коптев Константин):\n {'{:,}'.format(total).replace(',', ' ')} р.')

        file_path = 'БИГ ДЭЙТА.xlsx'
        with open(file_path, 'rb') as f:
            bot.send_document(message.chat.id, f)
        workbook = openpyxl.load_workbook('БИГ ДЭЙТА.xlsx')
        sheet = workbook['Лист1']
        total = 0
        total += sum(float(re.sub(r'р\.|\s|,00', '', row[6])) for row in sheet.iter_rows(min_row=2, values_only=True) if row[6] is not None)
        bot.send_message(chat_id, f'Сумма дебиторской задолженности (Общая): \n {'{:,}'.format(total).replace(',', ' ')} р.')

    else:
        bot.send_message(chat_id, "У вас нет доступа к этой команде.")

def Vlad(chat_id_admin_Vlad):
    file_path = 'Тогумбвев Владислав.xlsx'
    with open(file_path, 'rb') as f:
        bot.send_document(chat_id_admin_Vlad, f)
    workbook = openpyxl.load_workbook('Тогумбвев Владислав.xlsx')
    sheet = workbook['Sheet1']
    total = 0
    total += sum(float(row[6]) for row in sheet.iter_rows(min_row=2, values_only=True) if row[6] is not None)
    bot.send_message(chat_id_admin_Vlad,
                     f'Сумма дебиторской задолженности (Клиентов Тогумбвев Владислав):\n {'{:,}'.format(total).replace(',', ' ')} р.')

    file_path = 'Максим Стрелков.xlsx'
    with open(file_path, 'rb') as f:
        bot.send_document(chat_id_admin_Vlad, f)
    workbook = openpyxl.load_workbook('Максим Стрелков.xlsx')
    sheet = workbook['Sheet1']
    total = 0
    total += sum(float(row[6]) for row in sheet.iter_rows(min_row=2, values_only=True) if row[6] is not None)
    bot.send_message(chat_id_admin_Vlad,
                     f'Сумма дебиторской задолженности (Клиентов Максим Стрелков):\n {'{:,}'.format(total).replace(',', ' ')} р.')

    file_path = 'Кондратьев Павел.xlsx'
    with open(file_path, 'rb') as f:
        bot.send_document(chat_id_admin_Vlad, f)
    workbook = openpyxl.load_workbook('Кондратьев Павел.xlsx')
    sheet = workbook['Sheet1']
    total = 0
    total += sum(float(row[6]) for row in sheet.iter_rows(min_row=2, values_only=True) if row[6] is not None)
    bot.send_message(chat_id_admin_Vlad,
                     f'Сумма дебиторской задолженности (Клиентов Кондратьев Павел):\n {'{:,}'.format(total).replace(',', ' ')} р.')

    file_path = 'Коптев Константин.xlsx'
    with open(file_path, 'rb') as f:
        bot.send_document(chat_id_admin_Vlad, f)
    workbook = openpyxl.load_workbook('Коптев Константин.xlsx')
    sheet = workbook['Sheet1']
    total = 0
    total += sum(float(row[6]) for row in sheet.iter_rows(min_row=2, values_only=True) if row[6] is not None)
    bot.send_message(chat_id_admin_Vlad,
                     f'Сумма дебиторской задолженности (Клиентов Коптев Константин):\n {'{:,}'.format(total).replace(',', ' ')} р.')

    file_path = 'БИГ ДЭЙТА.xlsx'
    with open(file_path, 'rb') as f:
        bot.send_document(chat_id_admin_Vlad, f)
    workbook = openpyxl.load_workbook('БИГ ДЭЙТА.xlsx')
    sheet = workbook['Лист1']
    total = 0
    total += sum(float(re.sub(r'р\.|\s|,00', '', row[6])) for row in sheet.iter_rows(min_row=2, values_only=True) if
                row[6] is not None)
    bot.send_message(chat_id_admin_Vlad, f'Сумма дебиторской задолженности (Общая): \n {'{:,}'.format(total).replace(',', ' ')} р.')
def Zhenya(chat_id_admin_Zhenya):
    file_path = 'Тогумбвев Владислав.xlsx'
    with open(file_path, 'rb') as f:
        bot.send_document(chat_id_admin_Zhenya, f)
    workbook = openpyxl.load_workbook('Тогумбвев Владислав.xlsx')
    sheet = workbook['Sheet1']
    total = 0
    total += sum(float(row[6]) for row in sheet.iter_rows(min_row=2, values_only=True) if row[6] is not None)
    bot.send_message(chat_id_admin_Zhenya,
                     f'Сумма дебиторской задолженности (Клиентов Тогумбвев Владислав):\n {'{:,}'.format(total).replace(',', ' ')} р.')

    file_path = 'Максим Стрелков.xlsx'
    with open(file_path, 'rb') as f:
        bot.send_document(chat_id_admin_Zhenya, f)
    workbook = openpyxl.load_workbook('Максим Стрелков.xlsx')
    sheet = workbook['Sheet1']
    total = 0
    total += sum(float(row[6]) for row in sheet.iter_rows(min_row=2, values_only=True) if row[6] is not None)
    bot.send_message(chat_id_admin_Zhenya,
                     f'Сумма дебиторской задолженности (Клиентов Максим Стрелков):\n {'{:,}'.format(total).replace(',', ' ')} р.')

    file_path = 'Кондратьев Павел.xlsx'
    with open(file_path, 'rb') as f:
        bot.send_document(chat_id_admin_Zhenya, f)
    workbook = openpyxl.load_workbook('Кондратьев Павел.xlsx')
    sheet = workbook['Sheet1']
    total = 0
    total += sum(float(row[6]) for row in sheet.iter_rows(min_row=2, values_only=True) if row[6] is not None)
    bot.send_message(chat_id_admin_Zhenya,
                     f'Сумма дебиторской задолженности (Клиентов Кондратьев Павел):\n {'{:,}'.format(total).replace(',', ' ')} р.')

    file_path = 'Коптев Константин.xlsx'
    with open(file_path, 'rb') as f:
        bot.send_document(chat_id_admin_Zhenya, f)
    workbook = openpyxl.load_workbook('Коптев Константин.xlsx')
    sheet = workbook['Sheet1']
    total = 0
    total += sum(float(row[6]) for row in sheet.iter_rows(min_row=2, values_only=True) if row[6] is not None)
    bot.send_message(chat_id_admin_Zhenya,
                     f'Сумма дебиторской задолженности (Клиентов Коптев Константин):\n {'{:,}'.format(total).replace(',', ' ')} р.')

    file_path = 'БИГ ДЭЙТА.xlsx'
    with open(file_path, 'rb') as f:
        bot.send_document(chat_id_admin_Zhenya, f)
    workbook = openpyxl.load_workbook('БИГ ДЭЙТА.xlsx')
    sheet = workbook['Лист1']
    total = 0
    total += sum(float(re.sub(r'р\.|\s|,00', '', row[6])) for row in sheet.iter_rows(min_row=2, values_only=True) if
                row[6] is not None)
    bot.send_message(chat_id_admin_Zhenya, f'Сумма дебиторской задолженности (Общая): \n {'{:,}'.format(total).replace(',', ' ')} р.')
def Pasha(chat_id_manager_Pasha):
    file_path = 'Кондратьев Павел.xlsx'
    with open(file_path, 'rb') as f:
        bot.send_document(chat_id_manager_Pasha, f)

    workbook = openpyxl.load_workbook('Кондратьев Павел.xlsx')
    sheet = workbook['Sheet1']
    total = 0
    total += sum(float(row[6]) for row in sheet.iter_rows(min_row=2, values_only=True) if row[6] is not None)
    bot.send_message(chat_id_manager_Pasha, f'Сумма дебиторской задолженности:\n {'{:,}'.format(total).replace(',', ' ')} р.')
def Max(chat_id_manager_Max):
    file_path = 'Максим Стрелков.xlsx'
    with open(file_path, 'rb') as f:
        bot.send_document(chat_id_manager_Max, f)

    workbook = openpyxl.load_workbook('Максим Стрелков.xlsx')
    sheet = workbook['Sheet1']
    total = 0
    total += sum(float(row[6]) for row in sheet.iter_rows(min_row=2, values_only=True) if row[6] is not None)
    bot.send_message(chat_id_manager_Max, f'Сумма дебиторской задолженности:\n {'{:,}'.format(total).replace(',', ' ')} р.')
def Anton(chat_id_Anton):
    file_path = 'БИГ ДЭЙТА.xlsx'
    with open(file_path, 'rb') as f:
        bot.send_document(chat_id_Anton, f)
    workbook = openpyxl.load_workbook('БИГ ДЭЙТА.xlsx')
    sheet = workbook['Лист1']
    total = 0
    total += sum(float(re.sub(r'р\.|\s|,00', '', row[6])) for row in sheet.iter_rows(min_row=2, values_only=True) if
                 row[6] is not None)
    bot.send_message(chat_id_Anton,
                     f'Сумма дебиторской задолженности (Общая): \n {'{:,}'.format(total).replace(',', ' ')} р.')

def job():
    Vlad('1699179440')
    Zhenya('397255972')
    Max('6655526374')
    Pasha('6720840984')
    Anton("6707703501")

def run_bot():
    bot.infinity_polling()

def schedule_tasks():
    schedule.every().monday.at("10:30").do(job)
    while True:
        schedule.run_pending()
        time.sleep(1)

bot_thread = threading.Thread(target=run_bot)
scheduler_thread = threading.Thread(target=schedule_tasks)

bot_thread.start()
scheduler_thread.start()

try:
    while True:
        time.sleep(1)
except KeyboardInterrupt:
    pass

bot.stop_polling()
bot_thread.join()
scheduler_thread.join()